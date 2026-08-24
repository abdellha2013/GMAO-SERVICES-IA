"""
app/data_sources/file/xlsx_loader.py
====================================

Loader spécialisé pour les fichiers Microsoft Excel XLSX.

Ce loader :

- valide le fichier ;
- vérifie l'extension ;
- détecte le type MIME ;
- extrait le contenu de toutes les feuilles ;
- construit un SourceDocument standardisé.

Le résultat est ensuite utilisé par le pipeline RAG :

XLSX
   ↓
XLSXLoader
   ↓
SourceDocument
   ↓
Parser
   ↓
Chunker
   ↓
Embedding
"""

from __future__ import annotations

from pathlib import Path
from typing import Final

from app.data_sources.file.file_source import FileSource
from app.models.document import SourceDocument


class XLSXLoader(FileSource):
    """
    Loader pour les fichiers Microsoft Excel XLSX.

    Parameters
    ----------
    source:
        Chemin du fichier XLSX.
    """

    SUPPORTED_EXTENSIONS: Final[tuple[str, ...]] = (".xlsx",)

    def __init__(self, source: str | Path) -> None:
        """
        Initialise le loader XLSX.
        """

        super().__init__(source)

    # ==========================================================
    # Source Information
    # ==========================================================

    @property
    def source_name(self) -> str:
        """
        Retourne le nom du fichier source.
        """

        return self.filename

    # ==========================================================
    # Load
    # ==========================================================

    def load(self) -> SourceDocument:
        """
        Charge et extrait le contenu du fichier XLSX.

        Le contenu extrait comprend :

        - toutes les feuilles ;
        - toutes les lignes ;
        - toutes les cellules ;
        - les métadonnées du classeur.

        Returns
        -------
        SourceDocument
            Document standardisé.

        Raises
        ------
        FileValidationError
            Si le fichier n'est pas un XLSX valide.

        FileLoadingError
            Si le document ne peut pas être chargé.
        """

        from zipfile import BadZipFile

        from app.data_sources.file.validators import (
            ensure_non_empty_content,
            ensure_zip_based_format,
        )
        from app.exceptions import (
            FileLoadingError,
            InvalidXLSXError,
        )

        self.logger.info(
            "Loading XLSX file '%s'.",
            self.path,
        )

        # ==========================================================
        # Validation
        # ==========================================================

        self.validate()

        self.ensure_extension(
            *self.SUPPORTED_EXTENSIONS
        )

        ensure_zip_based_format(self.path)

        # ==========================================================
        # Import dependency
        # ==========================================================

        try:

            from openpyxl import load_workbook

        except ImportError as exc:

            raise FileLoadingError(
                message=(
                    "The 'openpyxl' package is required "
                    "to load XLSX files."
                ),
                original=exc,
            ) from exc

        # ==========================================================
        # Load workbook
        # ==========================================================

        try:

            workbook = load_workbook(
                filename=self.path,
                data_only=True,
                read_only=True,
            )

        except BadZipFile as exc:

            self.logger.error(
                "XLSX file '%s' is not a valid Microsoft Excel workbook.",
                self.path,
            )

            raise InvalidXLSXError(
                message=(
                    f"File '{self.filename}' is not a valid "
                    f"Microsoft Excel workbook."
                ),
                original=exc,
            ) from exc

        except Exception as exc:

            self.logger.error(
                "Unable to open XLSX file '%s'.",
                self.path,
            )

            raise FileLoadingError(
                message=(
                    f"Unable to load XLSX file "
                    f"'{self.filename}'."
                ),
                original=exc,
            ) from exc
        
        # ==========================================================
        # Extract worksheets
        # ==========================================================

        worksheets: list[str] = []

        sheet_count = 0
        row_count = 0
        cell_count = 0
        sheet_names: list[str] = []

        try:
            for worksheet in workbook.worksheets:

                sheet_count += 1

                rows: list[str] = []

                for row in worksheet.iter_rows(values_only=True):

                    values = [
                        str(value).strip()
                        for value in row
                        if value is not None and str(value).strip()
                    ]

                    if not values:
                        continue

                    row_count += 1
                    cell_count += len(values)

                    rows.append(
                        " | ".join(values)
                    )

                if rows:

                    worksheets.append(
                        "\n".join(
                            [
                                f"Worksheet: {worksheet.title}",
                                "-" * 40,
                                *rows,
                            ]
                        )
                    )

            # Copie des noms de feuille avant fermeture du classeur.
            sheet_names = list(workbook.sheetnames)

        finally:
            # En mode read_only, openpyxl garde un handle fichier
            # ouvert tant que close() n'est pas appelé explicitement.
            workbook.close()

        # ==========================================================
        # Build content
        # ==========================================================

        content = "\n\n".join(worksheets)

        # ==========================================================
        # Empty workbook
        # ==========================================================

        ensure_non_empty_content(
            content,
            self.filename,
            logger=self.logger,
        )

        # ==========================================================
        # Metadata
        # ==========================================================

        metadata = dict(
            self.metadata()
        )

        metadata.update(
            {
                "content_length": len(content),
                "worksheets_count": sheet_count,
                "rows_count": row_count,
                "cells_count": cell_count,
                "sheet_names": workbook.sheetnames,
            }
        )

        # ==========================================================
        # Result
        # ==========================================================

        self.logger.info(
            "XLSX file '%s' loaded successfully.",
            self.filename,
        )

        return SourceDocument(
            source_name=self.filename,
            source_type="xlsx",
            source_path=self.path,
            content=content,
            mime_type=self.mime_type,
            size=self.size,
            created_at=self.created_at,
            updated_at=self.modified_at,
            metadata=metadata,
        )